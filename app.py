from flask import Flask, render_template, request, redirect, session, Response
import sqlite3
import openpyxl
from io import BytesIO
from datetime import date

app = Flask(__name__)
# Esta "llave secreta" es obligatoria para que Flask pueda poner la pulserita VIP (session)
app.secret_key = "secreto_flores_2026" 

# --- CREACIÓN DE LAS TABLAS ---
def inicializar_bd():
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()

    # 1. Tabla de Obras abstractas
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Obras (
            mfn INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            autor TEXT NOT NULL,
            editorial TEXT,
            anio TEXT
        )
    ''')

    # 2. Tabla de Ejemplares físicos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Ejemplares (
            nro_inventario INTEGER PRIMARY KEY,
            mfn_vinculado INTEGER,
            signatura_topografica TEXT,
            observaciones TEXT,
            FOREIGN KEY (mfn_vinculado) REFERENCES Obras (mfn)
        )
    ''')

    # 3. Tabla de Socios (Vecinos)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Socios (
            id_socio INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre_completo TEXT NOT NULL,
            telefono TEXT NOT NULL
        )
    ''')

    # 4. Tabla de Préstamos (El registro histórico)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Prestamos (
            id_prestamo INTEGER PRIMARY KEY AUTOINCREMENT,
            id_socio INTEGER,
            nro_inventario INTEGER,
            fecha_prestamo DATE,
            fecha_devolucion DATE,
            FOREIGN KEY (id_socio) REFERENCES Socios (id_socio),
            FOREIGN KEY (nro_inventario) REFERENCES Ejemplares (nro_inventario)
        )
    ''')

    conexion.commit()
    conexion.close()

# Llamamos a la función para que cree las tablas ni bien arranca el programa
inicializar_bd()

def formatear_titulo(texto):
    if not texto:
        return ""
    # Separamos el texto por los puntos y espacios (ej: "Titulo. Subtitulo")
    partes = texto.strip().split('. ')
    # Usamos capitalize() que solo pone en mayúscula la primerísima letra de la frase
    partes_corregidas = [parte.capitalize() for parte in partes]
    # Volvemos a unir todo
    return '. '.join(partes_corregidas)

@app.route('/')
def inicio():
    # 1. Detectamos en qué página estamos (por defecto la 1)
    pagina = request.args.get('pagina', 1, type=int)
    libros_por_pagina = 10
    offset = (pagina - 1) * libros_por_pagina
    
    busqueda = request.args.get('q')
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    
    # Consulta base con el "Radar" de préstamos
    consulta_base = '''
        SELECT e.nro_inventario, o.titulo, o.autor, o.editorial, o.anio, e.signatura_topografica, e.observaciones, o.mfn,
               (SELECT id_prestamo FROM Prestamos p WHERE p.nro_inventario = e.nro_inventario AND p.fecha_devolucion IS NULL LIMIT 1) as estado_prestamo
        FROM Ejemplares e
        JOIN Obras o ON e.mfn_vinculado = o.mfn
    '''
    
    if busqueda:
        termino = f"%{busqueda}%"
        # Si hay búsqueda, mostramos todo lo que coincida (sin paginar para no complicar el filtro ahora)
        cursor.execute(consulta_base + " WHERE o.titulo LIKE ? OR o.autor LIKE ?", (termino, termino))
    else:
        # Si no hay búsqueda, limitamos a 10 resultados empezando desde el offset
        cursor.execute(consulta_base + " LIMIT ? OFFSET ?", (libros_por_pagina, offset))
        
    lista_libros = cursor.fetchall()

    # Contamos el total para saber si hay una página siguiente
    cursor.execute("SELECT COUNT(*) FROM Ejemplares")
    total_libros = cursor.fetchone()[0]
    tiene_siguiente = total_libros > (pagina * libros_por_pagina)
    
    conexion.close()
    
    return render_template('index.html', 
                           libros=lista_libros, 
                           busqueda=busqueda, 
                           pagina=pagina, 
                           tiene_siguiente=tiene_siguiente)

@app.route('/nuevo')
def nuevo_libro():
    return render_template('nuevo_libro.html')

@app.route('/guardar', methods=['POST'])
def guardar_libro():
    # ACÁ SACAMOS LA SEGURIDAD PARA QUE TODOS PUEDAN CARGAR
    nro_inventario = request.form['nro_inventario']
    titulo = formatear_titulo(request.form['titulo'])
    autor = request.form['autor'].strip().title()
    editorial = request.form['editorial'].strip().title() if request.form['editorial'] else ""
    anio = request.form['anio']
    signatura = request.form['signatura'].strip().upper()
    observaciones = request.form['observaciones']

    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()

    try:
        # 1. Chequeamos que el Nro de inventario no exista para evitar el error feo
        cursor.execute("SELECT nro_inventario FROM Ejemplares WHERE nro_inventario = ?", (nro_inventario,))
        if cursor.fetchone():
            conexion.close()
            return render_template('nuevo_libro.html', error=f"Error: El Nro. de Inventario {nro_inventario} ya está en uso. Elegí otro.")

        # 2. LÓGICA MFN: Buscamos si la obra ya existe comparando Título y Autor
        cursor.execute("SELECT mfn FROM Obras WHERE LOWER(titulo) = LOWER(?) AND LOWER(autor) = LOWER(?)", (titulo, autor))
        obra_existente = cursor.fetchone()

        if obra_existente:
            mfn_asignado = obra_existente[0] # Ya la tenemos, usamos su MFN
        else:
            # Es un libro nuevo, creamos la ficha de la Obra
            cursor.execute('''
                INSERT INTO Obras (titulo, autor, editorial, anio)
                VALUES (?, ?, ?, ?)
            ''', (titulo, autor, editorial, anio))
            mfn_asignado = cursor.lastrowid # Atrapamos el MFN que recién se creó

        # 3. Guardamos el ejemplar físico apuntando a ese MFN
        cursor.execute('''
            INSERT INTO Ejemplares (nro_inventario, mfn_vinculado, signatura_topografica, observaciones)
            VALUES (?, ?, ?, ?)
        ''', (nro_inventario, mfn_asignado, signatura, observaciones))
        
        conexion.commit()
        conexion.close()
        return redirect('/')
        
    except Exception as e:
        conexion.close()
        return render_template('nuevo_libro.html', error=f"Ocurrió un error: {str(e)}")

# --- RUTAS DE SEGURIDAD ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        clave_ingresada = request.form['clave']
        if clave_ingresada == 'flores123': 
            session['admin'] = True 
            return redirect('/')
        else:
            return render_template('login.html', error="Clave incorrecta. Intentá de nuevo.")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('admin', None) 
    return redirect('/')

@app.route('/borrar/<int:nro_inventario>')
def borrar_libro(nro_inventario):
    if not session.get('admin'):
        return redirect('/')

    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    
    # Buscamos a qué MFN pertenece antes de borrar el ejemplar
    cursor.execute("SELECT mfn_vinculado FROM Ejemplares WHERE nro_inventario = ?", (nro_inventario,))
    resultado = cursor.fetchone()
    
    if resultado:
        mfn = resultado[0]
        cursor.execute("DELETE FROM Ejemplares WHERE nro_inventario = ?", (nro_inventario,))
        
        # Limpieza: Si no quedan más ejemplares de esta obra, borramos la obra también
        cursor.execute("SELECT COUNT(*) FROM Ejemplares WHERE mfn_vinculado = ?", (mfn,))
        if cursor.fetchone()[0] == 0:
            cursor.execute("DELETE FROM Obras WHERE mfn = ?", (mfn,))
            
        conexion.commit()
        
    conexion.close()
    return redirect('/')

@app.route('/editar/<int:nro_inventario>')
def editar_libro(nro_inventario):
    if not session.get('admin'):
        return redirect('/')

    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    # Traemos los datos unidos para llenar el formulario
    cursor.execute('''
        SELECT e.nro_inventario, o.titulo, o.autor, o.editorial, o.anio, e.signatura_topografica, e.observaciones 
        FROM Ejemplares e
        JOIN Obras o ON e.mfn_vinculado = o.mfn
        WHERE e.nro_inventario = ?
    ''', (nro_inventario,))
    libro_a_editar = cursor.fetchone() 
    conexion.close()

    return render_template('editar_libro.html', libro=libro_a_editar)


@app.route('/actualizar/<int:nro_inventario>', methods=['POST'])
def actualizar_libro(nro_inventario):
    if not session.get('admin'):
        return redirect('/')

    titulo = formatear_titulo(request.form['titulo'])
    autor = request.form['autor'].strip().title()
    editorial = request.form['editorial'].strip().title() if request.form['editorial'] else ""
    anio = request.form['anio']
    signatura = request.form['signatura'].strip().upper()
    observaciones = request.form['observaciones']

    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    
    # Buscamos qué MFN le corresponde a este ejemplar
    cursor.execute("SELECT mfn_vinculado FROM Ejemplares WHERE nro_inventario = ?", (nro_inventario,))
    mfn = cursor.fetchone()[0]

    # Actualizamos los datos de la Obra (Esto corrige errores de tipeo para todos los ejemplares de este libro)
    cursor.execute('''
        UPDATE Obras 
        SET titulo = ?, autor = ?, editorial = ?, anio = ?
        WHERE mfn = ?
    ''', (titulo, autor, editorial, anio, mfn))
    
    # Actualizamos los datos propios de este ejemplar
    cursor.execute('''
        UPDATE Ejemplares 
        SET signatura_topografica = ?, observaciones = ?
        WHERE nro_inventario = ?
    ''', (signatura, observaciones, nro_inventario))
    
    conexion.commit()
    conexion.close()

    return redirect('/')

@app.route('/socios')
def lista_socios():
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    cursor.execute("SELECT id_socio, nombre_completo, telefono FROM Socios")
    socios = cursor.fetchall()
    conexion.close()
    return render_template('socios.html', socios=socios)

@app.route('/guardar_socio', methods=['POST'])
def guardar_socio():
    nombre = request.form['nombre'].strip().title() # Nombre con mayúsculas iniciales
    telefono = request.form['telefono'].strip()
    
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    cursor.execute("INSERT INTO Socios (nombre_completo, telefono) VALUES (?, ?)", (nombre, telefono))
    conexion.commit()
    conexion.close()
    return redirect('/socios')

@app.route('/prestar/<int:nro_inventario>')
def formulario_prestamo(nro_inventario):
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    # Traemos el libro específico
    cursor.execute("SELECT o.titulo, e.nro_inventario FROM Ejemplares e JOIN Obras o ON e.mfn_vinculado = o.mfn WHERE e.nro_inventario = ?", (nro_inventario,))
    libro = cursor.fetchone()
    # Traemos los socios ordenados alfabéticamente para el buscador
    cursor.execute("SELECT id_socio, nombre_completo FROM Socios ORDER BY nombre_completo ASC")
    socios = cursor.fetchall()
    conexion.close()
    return render_template('prestar.html', libro=libro, socios=socios)

@app.route('/registrar_prestamo', methods=['POST'])
def registrar_prestamo():
    id_socio = request.form['id_socio']
    nro_inv = request.form['nro_inventario']
    hoy = date.today().strftime("%d/%m/%Y") # Fecha con formato día/mes/año
    
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    cursor.execute("INSERT INTO Prestamos (id_socio, nro_inventario, fecha_prestamo) VALUES (?, ?, ?)", (id_socio, nro_inv, hoy))
    conexion.commit()
    conexion.close()
    return redirect('/')

@app.route('/devolver/<int:nro_inventario>')
def devolver_libro(nro_inventario):
    # Por seguridad, solo el admin puede procesar devoluciones
    if not session.get('admin'):
        return redirect('/')

    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    
    # 1. Buscamos el registro del préstamo que todavía está "abierto" (sin fecha de devolución)
    cursor.execute("SELECT id_prestamo FROM Prestamos WHERE nro_inventario = ? AND fecha_devolucion IS NULL LIMIT 1", (nro_inventario,))
    prestamo_abierto = cursor.fetchone()
    
    if prestamo_abierto:
        id_del_prestamo = prestamo_abierto[0]
        # 2. Conseguimos la fecha de hoy con formato argentino
        hoy_devolucion = date.today().strftime("%d/%m/%Y")
        
        # 3. Actualizamos la ficha del préstamo poniéndole la fecha de devolución
        cursor.execute("UPDATE Prestamos SET fecha_devolucion = ? WHERE id_prestamo = ?", (hoy_devolucion, id_del_prestamo))
        conexion.commit()
        
    conexion.close()
    
    # Volvemos a recargar la pantalla principal
    return redirect('/')

@app.route('/editar_socio/<int:id_socio>')
def vista_editar_socio(id_socio):
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    cursor.execute("SELECT id_socio, nombre_completo, telefono FROM Socios WHERE id_socio = ?", (id_socio,))
    socio = cursor.fetchone()
    conexion.close()
    return render_template('editar_socio.html', socio=socio)

@app.route('/actualizar_socio', methods=['POST'])
def actualizar_socio():
    id_s = request.form['id_socio']
    nom = request.form['nombre']
    tel = request.form['telefono']
    
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    cursor.execute("UPDATE Socios SET nombre_completo = ?, telefono = ? WHERE id_socio = ?", (nom, tel, id_s))
    conexion.commit()
    conexion.close()
    return redirect('/socios')

@app.route('/historial_socio/<int:id_socio>')
def historial_socio(id_socio):
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    
    # 1. Buscamos los datos del socio para el título
    cursor.execute("SELECT nombre_completo FROM Socios WHERE id_socio = ?", (id_socio,))
    nombre_socio = cursor.fetchone()[0]
    
    # 2. Buscamos su historial (Libros prestados y devueltos)
    cursor.execute('''
        SELECT o.titulo, o.autor, p.fecha_prestamo, p.fecha_devolucion, p.nro_inventario
        FROM Prestamos p
        JOIN Ejemplares e ON p.nro_inventario = e.nro_inventario
        JOIN Obras o ON e.mfn_vinculado = o.mfn
        WHERE p.id_socio = ?
        ORDER BY p.fecha_prestamo DESC
    ''', (id_socio,))
    
    historial = cursor.fetchall()
    conexion.close()
    
    return render_template('historial_socio.html', 
                           nombre=nombre_socio, 
                           historial=historial, 
                           id_s=id_socio)

@app.route('/historial_libro/<int:nro_inv>')
def historial_libro(nro_inv):
    conexion = sqlite3.connect("biblioteca.db")
    cursor = conexion.cursor()
    
    # Buscamos el título del libro para el encabezado
    cursor.execute('''
        SELECT o.titulo FROM Obras o 
        JOIN Ejemplares e ON o.mfn = e.mfn_vinculado 
        WHERE e.nro_inventario = ?
    ''', (nro_inv,))
    titulo = cursor.fetchone()[0]

    # Buscamos quiénes lo tuvieron
    cursor.execute('''
        SELECT s.nombre_completo, p.fecha_prestamo, p.fecha_devolucion
        FROM Prestamos p
        JOIN Socios s ON p.id_socio = s.id_socio
        WHERE p.nro_inventario = ?
        ORDER BY p.fecha_prestamo DESC
    ''', (nro_inv,))
    
    historial = cursor.fetchall()
    conexion.close()
    return render_template('historial_libro.html', titulo=titulo, historial=historial, nro_inv=nro_inv)


# ==========================================
# --- NUEVAS RUTAS DE EXPORTACIÓN A EXCEL ---
# ==========================================

@app.route('/exportar_libros')
def exportar_libros():
    conexion = sqlite3.connect('biblioteca.db')
    cursor = conexion.cursor()
    cursor.execute('''
        SELECT e.nro_inventario, o.mfn, o.titulo, o.autor, o.editorial, o.anio, e.signatura_topografica, e.observaciones 
        FROM Ejemplares e 
        JOIN Obras o ON e.mfn_vinculado = o.mfn
    ''')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    
    # Encabezados
    ws.append(['Nro_Inventario', 'MFN', 'Título', 'Autor', 'Editorial', 'Año', 'Signatura', 'Observaciones'])
    
    for row in cursor.fetchall():
        ws.append(row)
        
    conexion.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                    headers={"Content-disposition": "attachment; filename=catalogo_cacho_el_kadri.xlsx"})

@app.route('/exportar_socios')
def exportar_socios():
    conexion = sqlite3.connect('biblioteca.db')
    cursor = conexion.cursor()
    cursor.execute("SELECT id_socio, nombre_completo, telefono FROM Socios")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Socios"
    
    ws.append(['ID_Socio', 'Nombre_Completo', 'Teléfono'])
    
    for row in cursor.fetchall():
        ws.append(row)
        
    conexion.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                    headers={"Content-disposition": "attachment; filename=socios_flores_solidario.xlsx"})

@app.route('/exportar_prestamos')
def exportar_prestamos():
    conexion = sqlite3.connect('biblioteca.db')
    cursor = conexion.cursor()
    cursor.execute('''
        SELECT p.id_prestamo, e.nro_inventario, o.titulo, s.nombre_completo, p.fecha_prestamo, p.fecha_devolucion
        FROM Prestamos p
        JOIN Socios s ON p.id_socio = s.id_socio
        JOIN Ejemplares e ON p.nro_inventario = e.nro_inventario
        JOIN Obras o ON e.mfn_vinculado = o.mfn
    ''')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Préstamos"
    
    ws.append(['ID_Préstamo', 'Nro_Inventario', 'Título_Libro', 'Socio', 'Fecha_Préstamo', 'Fecha_Devolución'])
    
    for row in cursor.fetchall():
        ws.append(row)
        
    conexion.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                    headers={"Content-disposition": "attachment; filename=historial_prestamos.xlsx"})

# ==========================================


if __name__ == '__main__':
    # El host '0.0.0.0' le dice a Flask que escuche a cualquier dispositivo de tu red
    app.run(host='0.0.0.0', port=5000, debug=True)